import pandas as pd
from class_of_cards import Card, Hand, Shoe
from openpyxl import load_workbook





def find_blackjack_move(hand, dealer_card, true_count=0,Print =False):
    # Assuming Card is defined elsewhere with attributes suit and value
    strategy_sheets = pd.ExcelFile('Strategy1.xlsx')
    
    # Select the right sheet based on true count
    if true_count < -1:
        sheet_name = "-1"
    elif true_count > 4:
        sheet_name = "4"
    else:
        sheet_name = str(true_count)
    
    # Read the specific sheet and set the header to the 6th row (index 5)
    # Assuming you want to read from column G to the end of the sheet
    df = strategy_sheets.parse(sheet_name, header=5, usecols="G:Q")  # Adjust "X" as needed based on your last column

    type_of_hand = hand.type_of_hand()
    if Print:
        print("The hand is:",type_of_hand)
        print('The dealer has', dealer_card.value)

    if len(type_of_hand) < 3:
        type_of_hand = int(type_of_hand)

    index = df[df['Hand'] == type_of_hand].index[0]
    # print(index)  # This prints the index where 'Hand' matches 'type_of_hand'


    dealer_card_value = dealer_card.value
    if dealer_card_value == 11:
        dealer_card_value = 'A'
    else:
        dealer_card_value = int(dealer_card_value)


    
      # Access the value at the found index and specific dealer card column
    if dealer_card_value in df.columns:
        move = df.loc[index, dealer_card_value]
        if Print:
            print(f"The recommended move for {type_of_hand} against a dealer's {dealer_card_value} is: {move}")
    else:
        print(f"No column found for dealer's card value: {dealer_card_value}")
        
    return move


def find_move_test(hand, dealer_card, true_count=0,Print =False):
# Assuming Card is defined elsewhere with attributes suit and value
    
    strategy_sheets = pd.ExcelFile('test_table.xlsx')
    
    # Select the right sheet based on true count
    # print(f'the true count is:{true_count}')
    if true_count < -1:
        sheet_name = "-1"
    elif true_count > 3:
        sheet_name = "3"
    else:
        sheet_name = str(true_count)
    
    # Read the specific sheet and set the header to the 6th row (index 5)
    # Assuming you want to read from column G to the end of the sheet
    df = strategy_sheets.parse(sheet_name, header=5, usecols="G:Q")  # Adjust "X" as needed based on your last column

    type_of_hand = hand.type_of_hand()
    if Print:
        print("The hand is:",type_of_hand)
        print('The dealer has', dealer_card.value)

    if len(type_of_hand) < 3:
        type_of_hand = int(type_of_hand)

    if type_of_hand == 21:
        print('hand is already 21')
        exit()
    index = df[df['Hand'] == type_of_hand].index[0]
    # print(index)  # This prints the index where 'Hand' matches 'type_of_hand'

    
    dealer_card_value = dealer_card.value
    if dealer_card_value == 11:
        dealer_card_value = 'A'
    else:
        dealer_card_value = int(dealer_card_value)


    
      # Access the value at the found index and specific dealer card column
    if dealer_card_value in df.columns:
        move = df.loc[index, dealer_card_value]
        if Print:
            print(f"The recommended move for {type_of_hand} against a dealer's {dealer_card_value} is: {move}")
    else:
        print(f"No column found for dealer's card value: {dealer_card_value}")
        
    return move


excel_test_file = r'test_table.xlsx'


def write_move_to_excel(player_cards, dealer_card, move_to_write, true_count=0, excel_path=excel_test_file):
    # Load the workbook and sheet for writing
    player_hand = Hand(cards=player_cards)

    workbook = load_workbook(excel_path)
    sheet_name = str(true_count) if -5 < true_count < 5 else "-1" if true_count < -4 else "4"
    sheet = workbook[sheet_name]

    # Read the specific sheet using pandas for easy location of the hand type
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=5, usecols="G:Q")  # Adjust "G:Q" if needed

    type_of_hand = player_hand.type_of_hand()
    if isinstance(type_of_hand, str) and len(type_of_hand) < 3:
        type_of_hand = int(type_of_hand)  # Convert to integer if it's a number in string form

    if type_of_hand == 21:
        print('Hand is already 21')
        return

    # Find the row in the Excel file that matches the hand type
    try:
        row = df[df['Hand'] == type_of_hand].index[0] + 7  # Correct offset to match Excel's row numbering
        # print(f'Row is {row}')
    except IndexError:
        print("Hand type not found in the sheet.")
        return

    # Convert dealer card value to appropriate column header
    dealer_card_value = 'A' if dealer_card.value == 11 else str(dealer_card.value)
    

    # Find the appropriate column
    column = None
    for col in range(8, sheet.max_column + 1): # Assuming hand type starts from column G which is index 7
        
        if str(sheet.cell(row=6, column=col).value) == dealer_card_value:  # Header row is 6, corrected
            column = col
            break

    if not column:
        print(f"No column found for dealer's card value: {dealer_card_value}")
        return


    # Write the move to the found cell
    sheet.cell(row=row, column=column).value = move_to_write
    workbook.save(excel_path)
    # print(f"Move '{move_to_write}' written to row {row}, column {column}.")




if __name__ == '__main__':


    card1 = Card('Hearts', '2')
    card2 = Card('Hearts', '10')
    # card3 = Card('Hearts', '4')

    shoe = Shoe()

    dealers_card = Card('Hearts', '4')

    cards = [card1,card2]

    # hand = Hand(cards= cards,shoe=shoe)


    # find_move_test(hand= hand, dealer_card= dealers_card,true_count=3,Print=True)
    move_to_write = 'D'
    write_move_to_excel(player_cards= cards, dealer_card= dealers_card, move_to_write=move_to_write, true_count=0)
